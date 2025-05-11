import xml.etree.ElementTree as ET
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


# Функция для открытия проводника и выбора файла
def select_file(text):
    # Создаем скрытое главное окно
    root = tk.Tk()
    root.withdraw()  # Скрываем основное окно

    # Настраиваем параметры проводника
    file_path = filedialog.askopenfilename(
        title=f"Выберите XML-файл {text}",
        filetypes=[("XML файлы", "*.xml"), ("Все файлы", "*.*")]
    )

    # Возвращаем путь к выбранному файлу
    return file_path


def sbis_parsing(file):
    # Парсим XML-файл
    tree = ET.parse(file)
    root = tree.getroot()
    result = defaultdict(list)

    # Находим элемент <КнигаПрод>
    document = root.find('Документ')
    if document is not None:
        book_sales = document.find('КнигаПрод')

        # Перебираем все элементы <КнПродСтр> внутри <КнигаПрод>
        for sale in book_sales.findall('КнПродСтр'):
            tmp = dict()

            tmp[sale.get('НомСчФПрод')] = (sale.get('ДатаСчФПрод'), float(sale.get('СтоимПродСФ')))

            sv_pokup_ur = sale.find('СвПокуп/СведЮЛ')
            if sv_pokup_ur is not None:
                result[sv_pokup_ur.get('ИННЮЛ')].append(tmp)
                continue

            sv_pokup_ip = sale.find('СвПокуп/СведИП')
            if sv_pokup_ip is not None:
                result[sv_pokup_ip.get('ИННФЛ')].append(tmp)
                continue
            else:
                result['Без ИНН'].append(tmp)
    return result


def parsing_1c(file):
    # Парсим XML-файл
    tree = ET.parse(file)
    root = tree.getroot()
    result = defaultdict(list)

    # Находим элемент <Документ>
    document = root.find('Документ')
    if document is not None:
        for sale in document.findall('СвПродаж'):
            tmp = dict()

            tmp[sale.get('НомерСчФ')] = float(sale.get('СтТовУчНалРубКоп'))

            i1 = sale.get('ИННЮЛ')
            i2 = sale.get('ИННФЛ')
            i3 = 'Без ИНН'
            inn = list(filter(lambda x: x is not None, [i1, i2, i3]))[0]
            result[inn].append(tmp)

    return result


def get_sbis_data(sbis_dict):
    for k, v in sbis_dict.items():
        v1, v2 = v
        return k, v1, v2

def get_sum_1c(data, sf):
    for elem in data:
        if sf in elem:
            return elem[sf]
    else:
        return 0


def to_sum(number):
    return str(number).replace('.', ',')


def create_xlsx(file_sb, file_1c):
    data = [['ИНН', 'Номер СФ', 'Дата СФ', 'Сумма СБИС', 'Сумма 1С', 'Разница']]
    # Создаем строки для ТЧ
    for inn, values in file_sb.items():     # По ИНН получаем список словарей из файла СБИС
        data_1c = file_1c[inn]      # По ИНН получаем список словарей из файла 1С
        for value in values:
            sf, sf_date, sum_sbis = get_sbis_data(value)
            if data is not None:
                sum_1c = get_sum_1c(data_1c, sf)
            else:
                sum_1c = 0
            res = sum_sbis - sum_1c
            data.append([inn, sf, sf_date, sum_sbis, sum_1c, res])

    # Создаем новую книгу
    workbook = Workbook()
    sheet = workbook.active

    column_widths = [25, 18, 15, 20, 20, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = width

    # Заполняем данные в Excel
    for row in data:
        sheet.append(row)

    # Формируем имя файла с текущей датой и временем
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M")  # Формат: ГГГГ-ММ-ДД_ЧЧ-ММ
    file_name = f'Сверка СБИС и 1С_{current_time}.xlsx'

    # Сохраняем файл
    workbook.save(file_name)


file_sbis = sbis_parsing(select_file('СБИС'))
file_1c = parsing_1c(select_file('1С'))

create_xlsx(file_sbis, file_1c)
print('Работа программы завершена!')
